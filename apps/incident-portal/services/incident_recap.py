from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO
import math
import re
from typing import Iterable
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.incident_admin_store import get_dashboard_records
from services.incident_store import parse_dt

ACTIVE_STATUSES = {"OPEN", "IN_PROGRESS", "PENDING"}
TERMINAL_STATUSES = {"RESOLVED", "CLOSED"}


def category_label_from_ticket_code(ticket_code: str) -> str:
    prefix = str(ticket_code or "").strip().upper().split("-", 1)[0]
    return {
        "NET": "Network",
        "APP": "Application",
        "HW": "Hardware",
        "OTH": "Other",
        "INC": "Other",
    }.get(prefix, "Other")


def parse_duration_minutes(value) -> int | None:
    text = str(value or "").strip().lower()
    if not text or text == "-":
        return None
    total = 0
    matched = False
    for amount, unit in re.findall(r"(\d+)\s*([dhm])", text):
        matched = True
        qty = int(amount)
        if unit == "d":
            total += qty * 1440
        elif unit == "h":
            total += qty * 60
        elif unit == "m":
            total += qty
    if matched:
        return total
    if text.isdigit():
        return int(text)
    return None


def format_minutes(minutes) -> str:
    if minutes is None:
        return "-"
    minutes = max(int(minutes), 0)
    days, rem = divmod(minutes, 1440)
    hours, mins = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if mins or not parts:
        parts.append(f"{mins}m")
    return " ".join(parts)


def terminal_duration_minutes(row) -> int | None:
    status = str(row.get("status") or "").strip().upper()
    if status not in TERMINAL_STATUSES:
        return None
    created_dt = parse_dt(row.get("created_at"))
    if status == "CLOSED":
        end_dt = parse_dt(row.get("closed_at")) or parse_dt(row.get("update_terakhir")) or parse_dt(row.get("resolved_at"))
    else:
        end_dt = parse_dt(row.get("resolved_at")) or parse_dt(row.get("update_terakhir")) or parse_dt(row.get("closed_at"))
    if not created_dt or not end_dt:
        return None
    return max(0, int((end_dt - created_dt).total_seconds() / 60))


def build_latest_status_entry_times(history_rows):
    latest = {}
    for row in history_rows or []:
        ticket_code = str(row.get("ticket_code", "")).strip().upper()
        status_after = str(row.get("status_after", "")).strip().upper()
        update_dt = parse_dt(row.get("update_time"))
        if not ticket_code or status_after not in ACTIVE_STATUSES or not update_dt:
            continue
        key = (ticket_code, status_after)
        prev_dt = latest.get(key)
        if prev_dt is None or update_dt > prev_dt:
            latest[key] = update_dt
    return latest


def status_stagnation_minutes(row, now, latest_status_entry_times):
    ticket_code = str(row.get("ticket_code", "")).strip().upper()
    status = str(row.get("status", "")).strip().upper()
    start_dt = latest_status_entry_times.get((ticket_code, status))
    if not start_dt:
        start_dt = parse_dt(row.get("update_terakhir")) or parse_dt(row.get("created_at"))
    if not start_dt:
        return None
    return max(0, int((now - start_dt).total_seconds() / 60))


def sla_level(status: str, stagnation_minutes: int | None) -> str:
    status = str(status or "").upper()
    if stagnation_minutes is None:
        return "-"
    if status == "OPEN":
        if stagnation_minutes > 120:
            return "RED"
        if stagnation_minutes > 60:
            return "YELLOW"
    elif status == "IN_PROGRESS":
        if stagnation_minutes > 240:
            return "RED"
        if stagnation_minutes > 120:
            return "YELLOW"
    elif status == "PENDING":
        if stagnation_minutes > 1440:
            return "RED"
        if stagnation_minutes > 240:
            return "YELLOW"
    return "GREEN"


def normalize_record(row, now, latest_status_entry_times):
    record = dict(row)
    record["ticket_code"] = row.get("ticket_code") or "-"
    record["alias"] = row.get("alias") or "-"
    record["status"] = str(row.get("status") or "-").upper()
    record["lokasi"] = row.get("lokasi") or "-"
    record["masalah"] = row.get("masalah") or "-"
    record["pelapor"] = row.get("pelapor") or "-"
    record["petugas"] = row.get("ditangani_oleh") or "-"
    record["alasan_pending"] = row.get("alasan_pending") or "-"
    record["catatan_terakhir"] = row.get("catatan_terakhir") or "-"
    stored_durasi = row.get("durasi") or "-"
    computed_terminal_duration = terminal_duration_minutes(row)
    record["durasi"] = format_minutes(computed_terminal_duration) if computed_terminal_duration is not None else stored_durasi
    record["kategori"] = category_label_from_ticket_code(record["ticket_code"])
    record["created_dt"] = parse_dt(row.get("created_at"))
    record["update_dt"] = parse_dt(row.get("update_terakhir")) or parse_dt(row.get("closed_at")) or parse_dt(row.get("resolved_at"))
    record["created_at"] = row.get("created_at") or "-"
    record["update_terakhir"] = row.get("update_terakhir") or row.get("closed_at") or row.get("resolved_at") or "-"
    record["durasi_minutes"] = computed_terminal_duration if computed_terminal_duration is not None else parse_duration_minutes(row.get("durasi"))
    record["stagnation_minutes"] = status_stagnation_minutes(record, now, latest_status_entry_times)
    record["sla_level"] = sla_level(record["status"], record["stagnation_minutes"])
    return record


def filter_records(records, filters):
    start_date = filters.get("start_date")
    end_date = filters.get("end_date")
    status = str(filters.get("status") or "").strip().upper()
    category = str(filters.get("category") or "").strip().lower()
    petugas = str(filters.get("petugas") or "").strip().lower()
    lokasi = str(filters.get("lokasi") or "").strip().lower()

    filtered = []
    for row in records:
        created_dt = row.get("created_dt")
        if start_date and (not created_dt or created_dt.date() < start_date):
            continue
        if end_date and (not created_dt or created_dt.date() > end_date):
            continue
        if status and row.get("status") != status:
            continue
        if category and row.get("kategori", "").lower() != category:
            continue
        if petugas and petugas not in str(row.get("petugas") or "").lower():
            continue
        if lokasi and lokasi not in str(row.get("lokasi") or "").lower():
            continue
        filtered.append(row)
    return filtered


def parse_filters(args):
    start_raw = str(args.get("start_date") or "").strip()
    end_raw = str(args.get("end_date") or "").strip()
    today = datetime.now().date()
    end_date = datetime.strptime(end_raw, "%Y-%m-%d").date() if end_raw else today
    start_date = datetime.strptime(start_raw, "%Y-%m-%d").date() if start_raw else (end_date - timedelta(days=29))
    if start_date > end_date:
        raise ValueError("Tanggal awal tidak boleh lebih besar dari tanggal akhir.")
    return {
        "start_date": start_date,
        "end_date": end_date,
        "status": str(args.get("status") or "").strip().upper(),
        "category": str(args.get("category") or "").strip(),
        "petugas": str(args.get("petugas") or "").strip(),
        "lokasi": str(args.get("lokasi") or "").strip(),
    }


def default_filter_values():
    today = datetime.now().date()
    start_date = today - timedelta(days=29)
    return {
        "start_date": start_date.isoformat(),
        "end_date": today.isoformat(),
        "status": "",
        "category": "",
        "petugas": "",
        "lokasi": "",
        "report_type": "daftar_tiket",
        "format": "xlsx",
    }


def load_filtered_records(filters):
    active, archive, history_rows = get_dashboard_records()
    now = datetime.now()
    latest_status_entry_times = build_latest_status_entry_times(history_rows)
    records = [normalize_record(row, now, latest_status_entry_times) for row in (active + archive)]
    return filter_records(records, filters)


def build_status_summary(records):
    counts = Counter(row.get("status") for row in records)
    return [
        ["Total tiket", len(records)],
        ["OPEN", counts.get("OPEN", 0)],
        ["IN_PROGRESS", counts.get("IN_PROGRESS", 0)],
        ["PENDING", counts.get("PENDING", 0)],
        ["RESOLVED", counts.get("RESOLVED", 0)],
        ["CLOSED", counts.get("CLOSED", 0)],
    ]


def build_petugas_summary(records):
    grouped = defaultdict(list)
    for row in records:
        grouped[row.get("petugas") or "-"] .append(row)
    result = []
    for petugas, items in grouped.items():
        terminal = [row for row in items if row.get("status") in TERMINAL_STATUSES and row.get("durasi_minutes") is not None]
        avg_minutes = round(sum(row["durasi_minutes"] for row in terminal) / len(terminal)) if terminal else None
        counts = Counter(row.get("status") for row in items)
        result.append({
            "petugas": petugas,
            "total_tiket": len(items),
            "open": counts.get("OPEN", 0),
            "in_progress": counts.get("IN_PROGRESS", 0),
            "pending": counts.get("PENDING", 0),
            "resolved": counts.get("RESOLVED", 0),
            "closed": counts.get("CLOSED", 0),
            "avg_durasi": format_minutes(avg_minutes),
            "avg_durasi_minutes": avg_minutes if avg_minutes is not None else math.inf,
        })
    return sorted(result, key=lambda row: (-row["total_tiket"], row["avg_durasi_minutes"], row["petugas"]))


def build_lokasi_summary(records):
    grouped = defaultdict(list)
    for row in records:
        grouped[row.get("lokasi") or "-"] .append(row)
    result = []
    for lokasi, items in grouped.items():
        counts = Counter(row.get("status") for row in items)
        top_issue = Counter(row.get("masalah") or "-" for row in items).most_common(1)
        terminal = [row for row in items if row.get("status") in TERMINAL_STATUSES and row.get("durasi_minutes") is not None]
        avg_minutes = round(sum(row["durasi_minutes"] for row in terminal) / len(terminal)) if terminal else None
        result.append({
            "lokasi": lokasi,
            "jumlah_tiket": len(items),
            "open": counts.get("OPEN", 0),
            "in_progress": counts.get("IN_PROGRESS", 0),
            "pending": counts.get("PENDING", 0),
            "selesai": counts.get("RESOLVED", 0) + counts.get("CLOSED", 0),
            "avg_durasi": format_minutes(avg_minutes),
            "avg_durasi_minutes": avg_minutes if avg_minutes is not None else math.inf,
            "masalah_terbanyak": top_issue[0][0] if top_issue else "-",
        })
    return sorted(result, key=lambda row: (-row["jumlah_tiket"], row["avg_durasi_minutes"], row["lokasi"]))


def build_kategori_summary(records):
    grouped = defaultdict(list)
    for row in records:
        grouped[row.get("kategori") or "Other"].append(row)
    result = []
    total = max(len(records), 1)
    for kategori, items in grouped.items():
        counts = Counter(row.get("status") for row in items)
        terminal = [row for row in items if row.get("status") in TERMINAL_STATUSES and row.get("durasi_minutes") is not None]
        avg_minutes = round(sum(row["durasi_minutes"] for row in terminal) / len(terminal)) if terminal else None
        result.append({
            "kategori": kategori,
            "jumlah_tiket": len(items),
            "persentase": round((len(items) / total) * 100, 1),
            "open": counts.get("OPEN", 0),
            "in_progress": counts.get("IN_PROGRESS", 0),
            "pending": counts.get("PENDING", 0),
            "selesai": counts.get("RESOLVED", 0) + counts.get("CLOSED", 0),
            "avg_durasi": format_minutes(avg_minutes),
            "avg_durasi_minutes": avg_minutes if avg_minutes is not None else math.inf,
        })
    return sorted(result, key=lambda row: (-row["jumlah_tiket"], row["avg_durasi_minutes"], row["kategori"]))


def build_urgent_summary(records):
    urgent = []
    for row in records:
        if row.get("sla_level") not in {"YELLOW", "RED"} or row.get("status") not in ACTIVE_STATUSES:
            continue
        item = dict(row)
        item["stagnation_label"] = format_minutes(row.get("stagnation_minutes"))
        urgent.append(item)
    return sorted(urgent, key=lambda row: (0 if row.get("sla_level") == "RED" else 1, -(row.get("stagnation_minutes") or 0), row.get("alias") or ""))


def build_report_dataset(report_type: str, filters):
    records = load_filtered_records(filters)
    period_label = f"{filters['start_date'].strftime('%d-%m-%Y')} s.d. {filters['end_date'].strftime('%d-%m-%Y')}"
    base = {
        "report_type": report_type,
        "period_label": period_label,
        "generated_at": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        "filters": filters,
        "records": records,
        "status_summary": build_status_summary(records),
    }
    if report_type == "daftar_tiket":
        base["title"] = "Rekap Daftar Tiket"
    elif report_type == "per_petugas":
        base["title"] = "Rekap Per Petugas"
        base["petugas_rows"] = build_petugas_summary(records)
    elif report_type == "per_lokasi_kategori":
        base["title"] = "Rekap Per Lokasi dan Kategori"
        base["lokasi_rows"] = build_lokasi_summary(records)
        base["kategori_rows"] = build_kategori_summary(records)
    elif report_type == "eksekutif":
        base["title"] = "Rekap Eksekutif Insiden"
        base["petugas_rows"] = build_petugas_summary(records)[:5]
        base["lokasi_rows"] = build_lokasi_summary(records)[:5]
        base["kategori_rows"] = build_kategori_summary(records)[:5]
        base["urgent_rows"] = build_urgent_summary(records)[:10]
    else:
        raise ValueError("Jenis rekap tidak dikenali.")
    return base


def _write_sheet(ws, title, headers, rows):
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 36)


def export_excel(dataset) -> BytesIO:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    summary_ws = wb.create_sheet("Ringkasan")
    _write_sheet(summary_ws, "Ringkasan", ["Indikator", "Nilai"], dataset["status_summary"])
    summary_ws.insert_rows(1, 4)
    summary_ws["A1"] = dataset["title"]
    summary_ws["A2"] = "Periode"
    summary_ws["B2"] = dataset["period_label"]
    summary_ws["A3"] = "Dibuat"
    summary_ws["B3"] = dataset["generated_at"]
    summary_ws["A1"].font = Font(bold=True, size=14)

    if dataset["report_type"] == "daftar_tiket":
        _write_sheet(
            wb.create_sheet("Daftar Tiket"),
            "Daftar Tiket",
            ["Alias", "Kode Tiket", "Tanggal Masuk", "Update Terakhir", "Status", "Kategori", "Lokasi", "Masalah", "Pelapor", "Petugas", "Alasan Pending", "Durasi"],
            [
                [r["alias"], r["ticket_code"], r["created_at"], r["update_terakhir"], r["status"], r["kategori"], r["lokasi"], r["masalah"], r["pelapor"], r["petugas"], r["alasan_pending"], r["durasi"]]
                for r in dataset["records"]
            ],
        )
    elif dataset["report_type"] == "per_petugas":
        _write_sheet(
            wb.create_sheet("Per Petugas"),
            "Per Petugas",
            ["Petugas", "Total Tiket", "OPEN", "IN_PROGRESS", "PENDING", "RESOLVED", "CLOSED", "Rata-rata Durasi"],
            [
                [r["petugas"], r["total_tiket"], r["open"], r["in_progress"], r["pending"], r["resolved"], r["closed"], r["avg_durasi"]]
                for r in dataset.get("petugas_rows", [])
            ],
        )
    elif dataset["report_type"] == "per_lokasi_kategori":
        _write_sheet(
            wb.create_sheet("Per Lokasi"),
            "Per Lokasi",
            ["Lokasi", "Jumlah Tiket", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Rata-rata Durasi", "Masalah Terbanyak"],
            [
                [r["lokasi"], r["jumlah_tiket"], r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"], r["masalah_terbanyak"]]
                for r in dataset.get("lokasi_rows", [])
            ],
        )
        _write_sheet(
            wb.create_sheet("Per Kategori"),
            "Per Kategori",
            ["Kategori", "Jumlah Tiket", "Persentase", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Rata-rata Durasi"],
            [
                [r["kategori"], r["jumlah_tiket"], f"{r['persentase']}%", r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"]]
                for r in dataset.get("kategori_rows", [])
            ],
        )
    elif dataset["report_type"] == "eksekutif":
        _write_sheet(
            wb.create_sheet("Petugas"),
            "Petugas",
            ["Petugas", "Total Tiket", "OPEN", "IN_PROGRESS", "PENDING", "RESOLVED", "CLOSED", "Rata-rata Durasi"],
            [
                [r["petugas"], r["total_tiket"], r["open"], r["in_progress"], r["pending"], r["resolved"], r["closed"], r["avg_durasi"]]
                for r in dataset.get("petugas_rows", [])
            ],
        )
        _write_sheet(
            wb.create_sheet("Lokasi"),
            "Lokasi",
            ["Lokasi", "Jumlah Tiket", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Rata-rata Durasi", "Masalah Terbanyak"],
            [
                [r["lokasi"], r["jumlah_tiket"], r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"], r["masalah_terbanyak"]]
                for r in dataset.get("lokasi_rows", [])
            ],
        )
        _write_sheet(
            wb.create_sheet("Kategori"),
            "Kategori",
            ["Kategori", "Jumlah Tiket", "Persentase", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Rata-rata Durasi"],
            [
                [r["kategori"], r["jumlah_tiket"], f"{r['persentase']}%", r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"]]
                for r in dataset.get("kategori_rows", [])
            ],
        )
        _write_sheet(
            wb.create_sheet("Urgent"),
            "Urgent",
            ["Alias", "Kode Tiket", "SLA", "Status", "Lokasi", "Masalah", "Petugas", "Durasi Aktif"],
            [
                [r["alias"], r["ticket_code"], r["sla_level"], r["status"], r["lokasi"], r["masalah"], r["petugas"], r.get("stagnation_label") or format_minutes(r["stagnation_minutes"])]
                for r in dataset.get("urgent_rows", [])
            ],
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _pdf_styles():
    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "RecapSmall",
        parent=styles["BodyText"],
        fontSize=8.5,
        leading=10,
        wordWrap="CJK",
        splitLongWords=True,
    )
    return {
        "title": ParagraphStyle("RecapTitle", parent=styles["Heading1"], fontSize=16, leading=20, spaceAfter=10),
        "heading": ParagraphStyle("RecapHeading", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=10, spaceAfter=6),
        "normal": styles["BodyText"],
        "header": ParagraphStyle(
            "RecapHeader",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=10,
            textColor=colors.whitesmoke,
            alignment=0,
        ),
        "small": small,
        "small_wrap": ParagraphStyle(
            "RecapSmallWrap",
            parent=small,
            wordWrap="CJK",
            splitLongWords=True,
        ),
    }


def _wrap_words_for_pdf(value, max_chars: int) -> str:
    text = str(value or "-").strip()
    if not text or text == "-" or len(text) <= max_chars:
        return text or "-"
    words = text.split()
    if len(words) <= 1:
        return text
    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= max_chars:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return "\n".join(lines)


def _pdf_cell(value, style, *, bold: bool = False, hard_wrap_chars: int | None = None):
    raw_text = str(value or "-")
    if hard_wrap_chars and hard_wrap_chars > 0:
        raw_text = _wrap_words_for_pdf(raw_text, hard_wrap_chars)
    text = escape(raw_text).replace("\n", "<br/>")
    if bold:
        text = f"<b>{text}</b>"
    return Paragraph(text, style)


def _table(data, widths=None):
    tbl = Table(data, colWidths=widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return tbl


def export_pdf(dataset) -> BytesIO:
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=1.2 * cm, rightMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = _pdf_styles()
    story = [
        Paragraph(dataset["title"], styles["title"]),
        Paragraph(f"Periode: {dataset['period_label']}", styles["normal"]),
        Paragraph(f"Dibuat: {dataset['generated_at']}", styles["normal"]),
        Spacer(1, 10),
        Paragraph("Ringkasan Status", styles["heading"]),
        _table([["Indikator", "Nilai"], *dataset["status_summary"]], widths=[8 * cm, 5 * cm]),
    ]

    if dataset["report_type"] == "daftar_tiket":
        daftar_rows = [
            ["Alias", "Kode", "Masuk", "Status", "Kategori", "Lokasi", "Masalah", "Petugas", "Durasi"]
        ] + [
            [
                _pdf_cell(r["alias"], styles["small"]),
                _pdf_cell(r["ticket_code"], styles["small"]),
                _pdf_cell(r["created_at"], styles["small_wrap"], hard_wrap_chars=16),
                _pdf_cell(r["status"], styles["small"]),
                _pdf_cell(r["kategori"], styles["small"]),
                _pdf_cell(r["lokasi"], styles["small_wrap"], hard_wrap_chars=14),
                _pdf_cell(r["masalah"], styles["small"]),
                _pdf_cell(r["petugas"], styles["small"]),
                _pdf_cell(r["durasi"], styles["small"]),
            ]
            for r in dataset["records"]
        ]
        story += [
            Spacer(1, 12),
            Paragraph("Daftar Tiket", styles["heading"]),
            _table(
                daftar_rows,
                widths=[1.5*cm, 3.3*cm, 2.4*cm, 2.2*cm, 2.2*cm, 3.1*cm, 6.2*cm, 2.6*cm, 2.3*cm],
            ),
        ]
    elif dataset["report_type"] == "per_petugas":
        story += [
            Spacer(1, 12),
            Paragraph("Rekap Per Petugas", styles["heading"]),
            _table(
                [["Petugas", "Total", "OPEN", "IN_PROGRESS", "PENDING", "RESOLVED", "CLOSED", "Rata-rata Durasi"]] + [
                    [r["petugas"], r["total_tiket"], r["open"], r["in_progress"], r["pending"], r["resolved"], r["closed"], r["avg_durasi"]]
                    for r in dataset.get("petugas_rows", [])
                ],
                widths=[5*cm, 2*cm, 1.5*cm, 2.2*cm, 1.8*cm, 1.9*cm, 1.8*cm, 3.2*cm],
            ),
        ]
    elif dataset["report_type"] == "per_lokasi_kategori":
        story += [
            Spacer(1, 12),
            Paragraph("Rekap Per Lokasi", styles["heading"]),
            _table(
                [["Lokasi", "Jumlah", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Avg Durasi", "Masalah Terbanyak"]] + [
                    [r["lokasi"], r["jumlah_tiket"], r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"], r["masalah_terbanyak"]]
                    for r in dataset.get("lokasi_rows", [])
                ],
                widths=[4.2*cm, 1.5*cm, 1.4*cm, 2*cm, 1.6*cm, 1.6*cm, 2.8*cm, 6.7*cm],
            ),
            Spacer(1, 12),
            Paragraph("Rekap Per Kategori", styles["heading"]),
            _table(
                [["Kategori", "Jumlah", "Persentase", "OPEN", "IN_PROGRESS", "PENDING", "Selesai", "Avg Durasi"]] + [
                    [r["kategori"], r["jumlah_tiket"], f"{r['persentase']}%", r["open"], r["in_progress"], r["pending"], r["selesai"], r["avg_durasi"]]
                    for r in dataset.get("kategori_rows", [])
                ],
                widths=[3.8*cm, 1.7*cm, 2*cm, 1.4*cm, 2*cm, 1.6*cm, 1.6*cm, 2.8*cm],
            ),
        ]
    elif dataset["report_type"] == "eksekutif":
        story += [
            Spacer(1, 12),
            Paragraph("Top 5 Petugas", styles["heading"]),
            _table(
                [["Petugas", "Total", "Resolved", "Closed", "Avg Durasi"]] + [
                    [r["petugas"], r["total_tiket"], r["resolved"], r["closed"], r["avg_durasi"]]
                    for r in dataset.get("petugas_rows", [])
                ],
                widths=[7*cm, 2*cm, 2*cm, 2*cm, 3*cm],
            ),
            Spacer(1, 12),
            Paragraph("Top 5 Lokasi", styles["heading"]),
            _table(
                [["Lokasi", "Jumlah", "Avg Durasi", "Masalah Terbanyak"]] + [
                    [r["lokasi"], r["jumlah_tiket"], r["avg_durasi"], r["masalah_terbanyak"]]
                    for r in dataset.get("lokasi_rows", [])
                ],
                widths=[5*cm, 1.8*cm, 2.8*cm, 10.4*cm],
            ),
            Spacer(1, 12),
            Paragraph("Top 5 Kategori", styles["heading"]),
            _table(
                [["Kategori", "Jumlah", "Persentase", "Avg Durasi"]] + [
                    [r["kategori"], r["jumlah_tiket"], f"{r['persentase']}%", r["avg_durasi"]]
                    for r in dataset.get("kategori_rows", [])
                ],
                widths=[6.2*cm, 2.3*cm, 3*cm, 3*cm],
            ),
            Spacer(1, 12),
            Paragraph("Tiket Perlu Perhatian", styles["heading"]),
            _table(
                [["Alias", "Kode", "SLA", "Status", "Lokasi", "Masalah", "Petugas", "Durasi Aktif"]] + [
                    [r["alias"], r["ticket_code"], r["sla_level"], r["status"], r["lokasi"], r["masalah"], r["petugas"], r.get("stagnation_label") or format_minutes(r["stagnation_minutes"])]
                    for r in dataset.get("urgent_rows", [])
                ],
                widths=[1.8*cm, 3.6*cm, 1.5*cm, 2.2*cm, 3.2*cm, 6.5*cm, 3.2*cm, 2.6*cm],
            ),
        ]

    doc.build(story)
    output.seek(0)
    return output
