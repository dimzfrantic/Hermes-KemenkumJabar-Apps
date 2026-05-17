from openpyxl import load_workbook
from models import User

HEADER_MAP = {
    'nip': 'nip',
    'niplama': 'nip_lama',
    'nama': 'nama',
    'name': 'nama',
    'unit': 'unit',
    'bagian': 'unit',
    'satker': 'unit',
    'unit_kerja': 'unit',
    'nomor_hp': 'nomor_hp',
    'no_hp': 'nomor_hp',
    'hp': 'nomor_hp',
    'phone': 'nomor_hp',
    'telepon': 'nomor_hp',
    'no_telp': 'nomor_hp',
    'no_telepon': 'nomor_hp',
}

def normalize_header(value):
    return str(value or '').strip().lower().replace(' ', '_')

def import_employees_from_excel(db, file_path, default_password, role='employee'):
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError('File Excel kosong')

    raw_headers = [normalize_header(value) for value in rows[0]]
    mapped_headers = [HEADER_MAP.get(header, header) for header in raw_headers]
    required = {'nip', 'nama', 'unit'}
    if not required.issubset(set(mapped_headers)):
        raise ValueError('Kolom wajib minimal: nip, nama, unit')

    created = 0
    updated = 0
    skipped = 0

    for row in rows[1:]:
        payload = {mapped_headers[idx]: row[idx] for idx in range(min(len(mapped_headers), len(row)))}
        nip = ''.join(str(payload.get('nip') or '').split())
        nama = str(payload.get('nama') or '').strip()
        unit = str(payload.get('unit') or '').strip()
        nomor_hp = str(payload.get('nomor_hp') or '').strip()
        if not nip or not nama or not unit:
            skipped += 1
            continue

        user = User.query.filter_by(nip=nip).first()
        if user is None:
            user = User(
                nip=nip,
                full_name=nama,
                unit=unit,
                phone=nomor_hp or None,
                role=role,
                must_change_password=True,
                is_active_user=True,
            )
            user.set_password(default_password)
            db.session.add(user)
            created += 1
        else:
            user.full_name = nama
            user.unit = unit
            user.phone = nomor_hp or user.phone
            user.role = role if user.role != 'admin' else 'admin'
            updated += 1

    db.session.commit()
    return {'created': created, 'updated': updated, 'skipped': skipped}
