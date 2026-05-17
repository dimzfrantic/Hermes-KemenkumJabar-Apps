from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from datetime import datetime
from sqlalchemy import extract, func
from models import db, RapatHarmonisasi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

operator_bp = Blueprint("operator", __name__, url_prefix="/operator")


@operator_bp.before_request
@login_required
def require_login():
    if current_user.is_pimpinan:
        flash("Pimpinan tidak memiliki akses input data.", "warning")
        return redirect(url_for("dashboard.index"))


@operator_bp.route("/")
def index():
    tanggal_dari = request.args.get("tanggal_dari", "")
    tanggal_sampai = request.args.get("tanggal_sampai", "")
    query = RapatHarmonisasi.query

    if tanggal_dari:
        try:
            query = query.filter(
                RapatHarmonisasi.tanggal_rapat
                >= datetime.strptime(tanggal_dari, "%Y-%m-%d").date()
            )
        except ValueError:
            pass
    if tanggal_sampai:
        try:
            query = query.filter(
                RapatHarmonisasi.tanggal_rapat
                <= datetime.strptime(tanggal_sampai, "%Y-%m-%d").date()
            )
        except ValueError:
            pass

    page = request.args.get("page", 1, type=int)
    data = query.order_by(
        RapatHarmonisasi.tanggal_rapat.desc(), RapatHarmonisasi.id.desc()
    ).paginate(page=page, per_page=20, error_out=False)

    return render_template(
        "operator/index.html",
        data=data,
        tanggal_dari=tanggal_dari,
        tanggal_sampai=tanggal_sampai,
    )


@operator_bp.route("/tambah", methods=["GET", "POST"])
def tambah():
    if request.method == "POST":
        try:
            record = RapatHarmonisasi(
                tanggal_rapat=datetime.strptime(
                    request.form["tanggal_rapat"], "%Y-%m-%d"
                ).date(),
                daerah=request.form["daerah"],
                jenis=request.form["jenis"],
                program_pembentukan=request.form["program_pembentukan"],
                urusan_pemerintahan=request.form["urusan_pemerintahan"],
                nama_raperda=request.form["nama_raperda"],
                tim_kerja=request.form["tim_kerja"],
                hasil=request.form.get("hasil", "Selesai"),
                metode=request.form.get("metode", "Normal"),
                keterangan=request.form.get("keterangan", ""),
                created_by=current_user.id,
            )
            db.session.add(record)
            db.session.commit()
            flash("Data berhasil ditambahkan.", "success")
            return redirect(url_for("operator.index"))
        except Exception as e:
            db.session.rollback()
            flash(f"Gagal menambahkan data: {str(e)}", "danger")

    return render_template(
        "operator/form.html",
        record=None,
        daerah_choices=RapatHarmonisasi.DAERAH_CHOICES,
    )


@operator_bp.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):
    record = RapatHarmonisasi.query.get_or_404(id)

    if request.method == "POST":
        try:
            record.tanggal_rapat = datetime.strptime(
                request.form["tanggal_rapat"], "%Y-%m-%d"
            ).date()
            record.daerah = request.form["daerah"]
            record.jenis = request.form["jenis"]
            record.program_pembentukan = request.form["program_pembentukan"]
            record.urusan_pemerintahan = request.form["urusan_pemerintahan"]
            record.nama_raperda = request.form["nama_raperda"]
            record.tim_kerja = request.form["tim_kerja"]
            record.hasil = request.form.get("hasil", "Selesai")
            record.metode = request.form.get("metode", "Normal")
            record.keterangan = request.form.get("keterangan", "")

            db.session.commit()
            flash("Data berhasil diupdate.", "success")
            return redirect(url_for("operator.index"))
        except Exception as e:
            db.session.rollback()
            flash(f"Gagal mengupdate data: {str(e)}", "danger")

    return render_template(
        "operator/form.html",
        record=record,
        daerah_choices=RapatHarmonisasi.DAERAH_CHOICES,
    )


@operator_bp.route("/hapus/<int:id>", methods=["POST"])
def hapus(id):
    record = RapatHarmonisasi.query.get_or_404(id)
    try:
        db.session.delete(record)
        db.session.commit()
        flash("Data berhasil dihapus.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal menghapus data: {str(e)}", "danger")

    return redirect(url_for("operator.index"))


@operator_bp.route("/export", methods=["GET"])
def export_excel():
    """Export data rapat harmonisasi ke Excel."""
    tahun = request.args.get("tahun", str(datetime.now().year))
    try:
        tahun = int(tahun)
    except ValueError:
        flash("Tahun tidak valid.", "danger")
        return redirect(url_for("operator.index"))

    BULAN_NAMES = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    COL_HEADERS = [
        "No", "Tanggal Rapat Harmonisasi", "Daerah", "Jenis",
        "Program Pembentukan", "Urusan Pemerintahan",
        "Nama Raperda/ Raperkada", "Tim Kerja", "Hasil", "Metode"
    ]

    rekap_jenis = {jenis: [0] * 12 for jenis in RapatHarmonisasi.JENIS_CHOICES}
    rekap_daerah = {daerah: [0] * 12 for daerah in RapatHarmonisasi.DAERAH_CHOICES}

    for bulan_idx, bulan_name in enumerate(BULAN_NAMES):
        ws = wb.active if bulan_idx == 0 else wb.create_sheet()
        ws.title = bulan_name
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

        # Header row
        for col_idx, header in enumerate(COL_HEADERS, start=1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data
        records = (
            RapatHarmonisasi.query.filter(
                extract("year", RapatHarmonisasi.tanggal_rapat) == tahun,
                extract("month", RapatHarmonisasi.tanggal_rapat) == bulan_idx + 1
            )
            .order_by(RapatHarmonisasi.tanggal_rapat, RapatHarmonisasi.id)
            .all()
        )

        for row_idx, r in enumerate(records, start=7):
            ws.cell(row=row_idx, column=1, value=row_idx - 6).border = thin_border
            ws.cell(row=row_idx, column=2, value=r.tanggal_rapat).border = thin_border
            ws.cell(row=row_idx, column=2).number_format = 'DD-MMM-YYYY'
            ws.cell(row=row_idx, column=3, value=r.daerah).border = thin_border
            ws.cell(row=row_idx, column=4, value=r.jenis).border = thin_border
            ws.cell(row=row_idx, column=5, value=r.program_pembentukan).border = thin_border
            ws.cell(row=row_idx, column=6, value=r.urusan_pemerintahan).border = thin_border
            ws.cell(row=row_idx, column=6).alignment = wrap_align
            ws.cell(row=row_idx, column=7, value=r.nama_raperda).border = thin_border
            ws.cell(row=row_idx, column=7).alignment = wrap_align
            ws.cell(row=row_idx, column=8, value=r.tim_kerja).border = thin_border
            ws.cell(row=row_idx, column=9, value=r.hasil).border = thin_border
            ws.cell(row=row_idx, column=10, value=r.metode).border = thin_border

            # Count for rekapitulasi
            if r.jenis in rekap_jenis:
                rekap_jenis[r.jenis][bulan_idx] += 1
            if r.daerah in rekap_daerah:
                rekap_daerah[r.daerah][bulan_idx] += 1

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 45
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15

    # Rekapitulasi sheet
    ws = wb.create_sheet(title="Rekapitulasi")

    # Rekap Per Jenis
    ws.cell(row=1, column=1, value="REKAPITULASI PER JENIS").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)

    rekap_headers = ["No", "JENIS"] + BULAN_NAMES + ["TOTAL"]
    for col_idx, header in enumerate(rekap_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        if header == "JENIS":
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row_idx, jenis in enumerate(RapatHarmonisasi.JENIS_CHOICES, start=4):
        ws.cell(row=row_idx, column=1, value=row_idx - 3).border = thin_border
        ws.cell(row=row_idx, column=2, value=jenis).border = thin_border
        total = 0
        for bulan_idx in range(12):
            val = rekap_jenis[jenis][bulan_idx]
            ws.cell(row=row_idx, column=bulan_idx + 3, value=val if val > 0 else None).border = thin_border
            total += val
        ws.cell(row=row_idx, column=15, value=total).border = thin_border
        ws.cell(row=row_idx, column=15).font = Font(bold=True)

    # TOTAL row
    total_row = 4 + len(RapatHarmonisasi.JENIS_CHOICES)
    ws.cell(row=total_row, column=1, value="").border = thin_border
    ws.cell(row=total_row, column=2, value="TOTAL").border = thin_border
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    grand_total = 0
    for bulan_idx in range(12):
        col_total = sum(rekap_jenis[j][bulan_idx] for j in RapatHarmonisasi.JENIS_CHOICES)
        ws.cell(row=total_row, column=bulan_idx + 3, value=col_total if col_total > 0 else None).border = thin_border
        grand_total += col_total
    ws.cell(row=total_row, column=15, value=grand_total).border = thin_border
    ws.cell(row=total_row, column=15).font = Font(bold=True)

    # Rekap Per Daerah
    rekap_daerah_start = total_row + 3
    ws.cell(row=rekap_daerah_start, column=1, value="REKAPITULASI PER DAERAH").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=rekap_daerah_start, start_column=1, end_row=rekap_daerah_start, end_column=15)

    for col_idx, header in enumerate(rekap_headers, start=1):
        cell = ws.cell(row=rekap_daerah_start + 1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        if header == "JENIS":
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for row_idx, daerah in enumerate(RapatHarmonisasi.DAERAH_CHOICES, start=rekap_daerah_start + 2):
        ws.cell(row=row_idx, column=1, value=row_idx - (rekap_daerah_start + 1)).border = thin_border
        ws.cell(row=row_idx, column=2, value=daerah).border = thin_border
        total = 0
        for bulan_idx in range(12):
            val = rekap_daerah[daerah][bulan_idx]
            ws.cell(row=row_idx, column=bulan_idx + 3, value=val if val > 0 else None).border = thin_border
            total += val
        ws.cell(row=row_idx, column=15, value=total).border = thin_border
        ws.cell(row=row_idx, column=15).font = Font(bold=True)

    # TOTAL row per daerah
    total_row_daerah = rekap_daerah_start + 2 + len(RapatHarmonisasi.DAERAH_CHOICES)
    ws.cell(row=total_row_daerah, column=1, value="").border = thin_border
    ws.cell(row=total_row_daerah, column=2, value="TOTAL").border = thin_border
    ws.cell(row=total_row_daerah, column=2).font = Font(bold=True)
    grand_total_daerah = 0
    for bulan_idx in range(12):
        col_total = sum(rekap_daerah[d][bulan_idx] for d in RapatHarmonisasi.DAERAH_CHOICES)
        ws.cell(row=total_row_daerah, column=bulan_idx + 3, value=col_total if col_total > 0 else None).border = thin_border
        grand_total_daerah += col_total
    ws.cell(row=total_row_daerah, column=15, value=grand_total_daerah).border = thin_border
    ws.cell(row=total_row_daerah, column=15).font = Font(bold=True)

    # Column widths for Rekapitulasi
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col_letter].width = 12

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Laporan_Rapat_Harmonisasi_{tahun}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
