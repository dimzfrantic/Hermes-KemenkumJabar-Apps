from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class ParsedWorkbook:
    sheet_names: list[str]
    selected_sheet: str
    headers: list[str]
    rows: list[dict]


class WorkbookValidationError(RuntimeError):
    pass


def load_participants(workbook_path: str, sheet_name: str | None = None) -> ParsedWorkbook:
    wb = load_workbook(filename=Path(workbook_path), read_only=True, data_only=True)
    available = wb.sheetnames
    chosen_name = sheet_name or available[0]
    if chosen_name not in available:
        raise WorkbookValidationError(f'Sheet {chosen_name} tidak ditemukan.')
    ws = wb[chosen_name]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        raise WorkbookValidationError('File Excel tidak memiliki data.')
    headers = [str(cell).strip() if cell is not None else '' for cell in raw_rows[0]]
    if not any(headers):
        raise WorkbookValidationError('Header kolom pada baris pertama kosong.')

    normalized_headers = []
    seen = {}
    for index, header in enumerate(headers, start=1):
        value = header or f'Kolom {index}'
        if value in seen:
            seen[value] += 1
            value = f'{value} ({seen[value]})'
        else:
            seen[value] = 1
        normalized_headers.append(value)

    rows = []
    for row_number, row_values in enumerate(raw_rows[1:], start=2):
        row_dict = {normalized_headers[i]: ('' if value is None else str(value).strip()) for i, value in enumerate(row_values[:len(normalized_headers)])}
        if not any(v for v in row_dict.values()):
            continue
        row_dict['_row_number'] = row_number
        rows.append(row_dict)

    if not rows:
        raise WorkbookValidationError('Tidak ada data peserta yang dapat diproses.')

    return ParsedWorkbook(
        sheet_names=available,
        selected_sheet=chosen_name,
        headers=normalized_headers,
        rows=rows,
    )


def validate_required_columns(parsed: ParsedWorkbook, required_columns: list[str]) -> list[str]:
    missing = [col for col in required_columns if col not in parsed.headers]
    return missing
